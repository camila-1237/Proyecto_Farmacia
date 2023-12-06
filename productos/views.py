from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from .forms import ProductoForm
from .models import Producto
from .models import TipoProducto
from datetime import datetime
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth import logout
from django.utils import timezone
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from .models import Venta, Transaccion
from .forms import VentaForm
from django.contrib.auth.models import User
from .models import Producto, Venta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone


from django.http import HttpResponse, FileResponse



def cerrar_sesion(request):
    logout(request)
    return redirect('login')

def index_view(request):
    return render(request, 'index.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
          
            return redirect('home')  
        else:
           
            pass

    return render(request, 'registration/login.html')


def registro_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']

        if password1 == password2:
            if User.objects.filter(username=username).exists():
                messages.error(request, 'El nombre de usuario ya está en uso.')
            elif User.objects.filter(email=email).exists():
                messages.error(request, 'El correo electrónico ya está registrado.')
            else:
                user = User.objects.create_user(username=username, email=email, password=password1)
                user = authenticate(username=username, password=password1)
                if user is not None:
                    login(request, user)
                    return redirect('login') 
        else:
            messages.error(request, 'Las contraseñas no coinciden.')

    return render(request, 'registration/registro.html')

def index( request ):
    return render(request, 'index.html')

def base( request ):
    return render(request, 'base.html')

@login_required
def listar_productos(request):
    productos = Producto.objects.all()
    return render(request, 'productos/listar_productos.html', {'productos': productos})

@login_required
def detalle_producto(request, producto_id):
    producto = get_object_or_404(Producto, id_producto=producto_id)
    return render(request, 'productos/detalle_producto.html', {'producto': producto})

from datetime import datetime
@login_required
def registrar_producto(request):
    tipos = TipoProducto.objects.all()

    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoForm()

    hoy = datetime.today().strftime('%Y-%m-%d')

    return render(request, 'productos/registrar_producto.html', {'form': form, 'tipos': tipos, 'hoy': hoy})
def modificar_producto(request, id):
    producto = get_object_or_404(Producto, id_producto=id)
    tipos = TipoProducto.objects.all()
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoForm(instance=producto)
        
    return render(request, 'productos/modificar_producto.html', {'form': form, 'tipos': tipos})


def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id_producto=producto_id)
    if request.method == 'POST':
        producto.delete()
        return redirect('listar_productos')

    return render(request, 'productos/eliminar_productos.html', {'producto': producto})

#JuanJo

def buscar_productos(request):
    if request.method == 'GET':
        consulta = request.GET.get('consulta')
        resultados = Producto.objects.filter(
            Q(nombre__icontains=consulta) |  
            Q(codigo_barras__icontains=consulta)  
            
        )

    return render(request, 'productos/buscador.html', {'resultados': resultados})


def alertas_bajo_inventario(request):
    if request.method == 'GET':
        
        productos_inventario_bajo = Producto.objects.filter(
            cantidad_inventario__lte=10
        )

        fecha_actual = timezone.now()
        fecha_limite = fecha_actual + timezone.timedelta(days=3)
        productos_proximos_a_vencer = Producto.objects.filter(
            Q(fecha_caducidad__gte=fecha_actual) & Q(fecha_caducidad__lte=fecha_limite)
        )

    return render(request, 'productos/alertas.html', {
        'productos_inventario_bajo': productos_inventario_bajo,
        'productos_proximos_a_vencer': productos_proximos_a_vencer,
    })


def venta(request):
   productos = Producto.objects.all()
   return render(request, 'venta/venta.html', {'productos': productos})


def listar_ventas(request):
    ventas = Venta.objects.all()

    listadoVentas = []

    for venta in ventas:

        productos = json.loads(venta.productos)
        listado_productos = []

        for p in productos:

            listado_productos.append({

                "nombre": p["nombre"],
                "cantidad": p["cantidad"]

            })

        listadoVentas.append({

            "id": venta.id_venta,
            "fecha": venta.fecha_venta,
            "estado": venta.finalizada,
            "total": venta.total,
            "productos": listado_productos

        })

    return render(request, 'venta/listar_ventas.html', {'ventas': listadoVentas})


@csrf_exempt
def crear_venta(request):
    try:
        total = float(request.POST.get('total'))
        productos_json = request.POST.get('productos')
        productos = json.loads(productos_json)

        Venta.objects.create(
            finalizada=True,
            total=total,
            productos=productos_json,
        )

        for producto in productos:
            prd = Producto.objects.get(id_producto=producto["id"])
            prd.cantidad_inventario -= int(producto["cantidad"])
            prd.save()

        return JsonResponse({"mensaje": "Exito"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
def informe_ventas(request):
    if request.method == 'POST':
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_fin_str = request.POST.get('fecha_fin')

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%d/%m/%Y').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%d/%m/%Y').date()
        except ValueError:
            return HttpResponse("Formato de fecha incorrecto. Utiliza el formato dd/mm/yyyy.")

        if fecha_inicio == fecha_fin:
            ventas = Venta.objects.filter(fecha_venta__date=fecha_inicio, finalizada=True)
        else:
            ventas = Venta.objects.filter(fecha_venta__range=[fecha_inicio, fecha_fin], finalizada=True)
    else:
        ventas = Venta.objects.filter(finalizada=True)

    total_general = 0
    productos_vendidos = {}

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Informe de Ventas"

    headers = ["ID Venta", "Total", "Productos Vendidos"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f"{col_letter}1"] = header
        worksheet[f"{col_letter}1"].font = Font(bold=True)

    for row_num, venta in enumerate(ventas, 2):
        productos = json.loads(venta.productos)

        for p in productos:
            nombre_producto = p["nombre"]
            cantidad_vendida = p["cantidad"]

            producto = Producto.objects.get(nombre=nombre_producto)
            precio_producto = producto.precio_venta

            total_producto = precio_producto * cantidad_vendida

            total_general += total_producto

            if nombre_producto in productos_vendidos:
                productos_vendidos[nombre_producto] += cantidad_vendida
            else:
                productos_vendidos[nombre_producto] = cantidad_vendida

        fecha_venta_excel = venta.fecha_venta.astimezone(timezone.utc).replace(tzinfo=None)

        worksheet.append([venta.id_venta, venta.total, json.dumps(productos)])

    excel_file_path = "informe_ventas.xlsx"
    workbook.save(excel_file_path)

    context = {'total_general': total_general, 'productos_vendidos': productos_vendidos, 'excel_file_path': excel_file_path}

    return render(request, 'venta/informe_ventas.html', context)


def descargar_excel(request):
    excel_file_path = "informe_ventas.xlsx"

    response = FileResponse(open(excel_file_path, 'rb'))
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = f'attachment; filename="{excel_file_path}"'

    return response
